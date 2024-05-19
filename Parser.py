from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
import time
def parse():
    url = 'https://omsk.cian.ru/kupit-kvartiru-1-komn-ili-2-komn/'
    page = requests.get(url)
    print(page.status_code)
    soup = BeautifulSoup(page.text, "html.parser")
    block = soup.findAll('a', class_='_93444fe79c--media--9P6wN')
    wb = Workbook()
    ws = wb.active
    for idx, link in enumerate(block, start=1):
        time.sleep(2)
        href = link.get('href')
        page2=requests.get(href)
        time.sleep(2)
        soup2=BeautifulSoup(page2.text,"html.parser")
        block2=soup2.findAll('div', class_='a10a3f92e9--container--pWxZo')
        print(page.status_code, page2.status_code)
        for data in block2:
            if data.find('h1'):
                des=data.text
                ws.cell(row=idx, column=7, value=des)
                print(des)
            break
        ws.cell(row=idx, column=1, value=href)
    wb.save('links.xlsx')
    print("Файл успешно создан")
